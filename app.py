import streamlit as st
import openpyxl
from datetime import datetime
import requests
import base64
import io

# Config from Streamlit Secrets
TOKEN = st.secrets["GITHUB_TOKEN"]
USER = st.secrets["GITHUB_USER"]
REPO = st.secrets["GITHUB_REPO"]
FILE_NAME = "GuangFaBank Transactions.xlsx"

st.set_page_config(page_title="Safe Finance Sync", page_icon="🔐")
st.title("📲 Safe Transaction Entry")
st.info("This version preserves your Formulas, Formatting, and all Sheets.")

with st.form("transaction_form", clear_on_submit=True):
    col1, col2 = st.columns(2)
    with col1:
        # User selects the table/sheet
        table_choice = st.selectbox("Target Sheet", ["Sheet2", "Sheet3"])
        date_val = st.date_input("Date", datetime.now())
    with col2:
        entity = st.selectbox("Entity", ["MV", "YEONG"])
        amount = st.number_input("Amount", min_value=0.0, format="%.2f")
    remarks = st.text_input("Remarks")
    submit = st.form_submit_button("Save & Sync")

if submit:
    try:
# 1. LOAD the existing workbook
        wb = openpyxl.load_workbook(FILE_NAME)
        
        if table_choice in wb.sheetnames:
            sheet = wb[table_choice]
            
            # Find the first truly empty row
            row_to_fill = 1
            while sheet.cell(row=row_to_fill, column=1).value is not None:
                row_to_fill += 1

            # --- TABLE EXPANSION LOGIC ---
            # Define the table name based on your selection
            target_table_name = "Table2" if table_choice == "Sheet2" else "Table3"
            
            # Find the table object within the sheet
            if target_table_name in sheet.tables:
                table = sheet.tables[target_table_name]
                
                # Get the current range (e.g., 'A1:D10')
                current_ref = table.ref 
                # Split it to get the start (A1) and end column (D)
                start_part, end_part = current_ref.split(':')
                
                # Create a new range ending at our new row (e.g., 'A1:D11')
                # We keep the column (D) and just change the number to row_to_fill
                import re
                end_column = re.sub(r'\d+', '', end_part) # Extracts "D"
                new_ref = f"{start_part}:{end_column}{row_to_fill}"
                
                # Update the table boundary
                table.ref = new_ref
            # -----------------------------

            # Insert Data
            sheet.cell(row=row_to_fill, column=1).value = date_val.strftime("%d-%m-%y")
            sheet.cell(row=row_to_fill, column=2).value = amount
            sheet.cell(row=row_to_fill, column=3).value = entity
            sheet.cell(row=row_to_fill, column=4).value = remarks
            
            # ... (Rest of your Save & Sync code)
            
            
            # Save the workbook to a temporary "buffer"
            buffer = io.BytesIO()
            wb.save(buffer)
            content_to_upload = buffer.getvalue()
            
            # 2. Sync back to GitHub
            url = f"https://api.github.com/repos/{USER}/{REPO}/contents/{FILE_NAME}"
            current_file = requests.get(url, headers={"Authorization": f"token {TOKEN}"}).json()
            sha = current_file['sha']

            # Encode and Push
            encoded_content = base64.b64encode(content_to_upload).decode("utf-8")
            data = {
                "message": f"Added row to {table_choice}",
                "content": encoded_content,
                "sha": sha
            }

            response = requests.put(url, json=data, headers={"Authorization": f"token {TOKEN}"})

            if response.status_code == 200:
                # IMPORTANT: Also save a local copy for the session
                with open(FILE_NAME, "wb") as f:
                    f.write(content_to_upload)
                st.success(f"✅ Successfully updated {table_choice} without losing formatting!")
            else:
                st.error(f"GitHub Sync Error: {response.json()}")
        else:
            st.error(f"Error: {table_choice} not found in Excel file.")

    except Exception as e:
        st.error(f"Error: {e}")


